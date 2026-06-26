import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime
import pandas as pd
import hashlib

# Nama file database Excel harian
DB_FILE = "Database_Multiuser.xlsx"

def hash_password(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

def inisialisasi_database():
    if not os.path.exists(DB_FILE):
        wb = openpyxl.Workbook()
        
        # Sheet 1: Data Akun User
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["username", "password"])
        
        # Sheet 2: Data Pengeluaran Semua User
        ws_data = wb.create_sheet(title="Pengeluaran")
        headers = ["Username", "Hari", "Tanggal", "Bulan", "Nama Barang / Kebutuhan", "Harga (Rp)"]
        ws_data.append(headers)
        
        # Styling Header Pengeluaran
        dark_emerald = "1B4D3E"
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color=dark_emerald, end_color=dark_emerald, fill_type="solid")
        
        for col in range(1, 7):
            cell = ws_data.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
            
        wb.save(DB_FILE)

inisialisasi_database()

# --- PENGATURAN HALAMAN WEB ---
st.set_page_config(page_title="Aplikasi Keuangan Bersama", page_icon="💰", layout="centered")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = ""

# --- HALAMAN BELUM LOGIN ---
if not st.session_state.logged_in:
    st.title("🔐 Akses Sistem Keuangan")
    tab_login, tab_reg = st.tabs(["🔑 Login", "📝 Daftar Akun Baru"])
    
    with tab_login:
        st.subheader("Silakan Login")
        user_input = st.text_input("Username", key="login_user").strip().lower()
        pass_input = st.text_input("Password", type="password", key="login_pass")
        
        if st.button("Masuk", type="primary"):
            if user_input and pass_input:
                wb = openpyxl.load_workbook(DB_FILE)
                ws = wb["Users"]
                user_found = False
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == user_input and ws.cell(row=row, column=2).value == hash_password(pass_input):
                        user_found = True
                        break
                if user_found:
                    st.session_state.logged_in = True
                    st.session_state.username = user_input
                    st.success(f"Selamat datang kembali, {user_input.capitalize()}!")
                    st.rerun()
                else:
                    st.error("Username atau Password salah!")
            else:
                st.warning("Semua kolom harus diisi!")

    with tab_reg:
        st.subheader("Buat Akun Baru")
        new_user = st.text_input("Buat Username", key="reg_user").strip().lower()
        new_pass = st.text_input("Buat Password", type="password", key="reg_pass")
        confirm_pass = st.text_input("Konfirmasi Password", type="password", key="reg_pass_conf")
        
        if st.button("Daftar Sekarang"):
            if new_user and new_pass and confirm_pass:
                if new_pass != confirm_pass:
                    st.error("Konfirmasi password tidak cocok!")
                else:
                    wb = openpyxl.load_workbook(DB_FILE)
                    ws = wb["Users"]
                    username_exists = False
                    for row in range(2, ws.max_row + 1):
                        if ws.cell(row=row, column=1).value == new_user:
                            username_exists = True
                            break
                    if username_exists:
                        st.error("Username sudah terpakai!")
                    else:
                        ws.append([new_user, hash_password(new_pass)])
                        wb.save(DB_FILE)
                        st.success("Akun berhasil dibuat! Silakan Login.")
            else:
                st.warning("Semua kolom wajib diisi!")

# --- HALAMAN SETELAH BERHASIL LOGIN ---
else:
    st.sidebar.title(f"👤 Akun: {st.session_state.username.capitalize()}")
    if st.sidebar.button("🚪 Logout/Keluar"):
        st.session_state.logged_in = False
        st.session_state.username = ""
        st.rerun()
        
    st.title(f"💰 Dompet Digital: {st.session_state.username.capitalize()}")
    st.write("Catat pengeluaran pribadi Anda di bawah ini.")
    st.markdown("---")
    
    # Form Input Data
    nama_barang = st.text_input("Nama Barang / Kebutuhan", placeholder="Contoh: Nasi Goreng, Pulsa")
    harga = st.number_input("Harga (Rp)", min_value=0, step=1000, value=0)
    
    if st.button("Simpan Pengeluaran", type="primary"):
        if nama_barang == "":
            st.error("Nama barang tidak boleh kosong!")
        elif harga <= 0:
            st.error("Harga harus lebih dari 0!")
        else:
            sekarang = datetime.now()
            hari_indo = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
            hari = hari_indo[sekarang.weekday()]
            tanggal = hardcore_date = sekarang.strftime("%Y-%m-%d")
            bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                          "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
            bulan = bulan_indo[sekarang.month - 1]
            
            wb = openpyxl.load_workbook(DB_FILE)
            ws = wb["Pengeluaran"]
            ws.append([st.session_state.username, hari, tanggal, bulan, nama_barang, harga])
            ws.cell(row=ws.max_row, column=6).number_format = '#,##0'
            wb.save(DB_FILE)
            st.success(f"Berhasil disimpan: {nama_barang}")
            st.rerun()

    st.markdown("---")
    
    # --- FITUR BARU: MANAGEMENT DATA & HAPUS DATA SALAH INPUT ---
    st.subheader("📊 Riwayat & Kelola Pengeluaran")
    
    try:
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb["Pengeluaran"]
        
        # Ambil data dari Excel ke list untuk diproses
        data_rows = []
        for row in range(2, ws.max_row + 1):
            vals = [ws.cell(row=row, column=col).value for col in range(1, 7)]
            # Menyimpan info nomor baris asli Excel di akhir elemen list (indeks ke-6)
            vals.append(row) 
            data_rows.append(vals)
            
        # Filter hanya data milik user yang sedang aktif login
        user_entries = [r for r in data_rows if r[0] == st.session_state.username]
        
        if user_entries:
            # Tampilkan tombol hapus per baris data
            st.write("Jika ada data yang salah input, klik tombol **Hapus** di samping kanan data:")
            
            total_user = 0
            for entry in user_entries:
                # entry format: [username, hari, tanggal, bulan, nama_barang, harga, baris_excel]
                row_id = entry[6]
                nama_b = entry[4]
                tgl_b = entry[2]
                harga_b = entry[5]
                total_user += harga_b
                
                # Buat layout kolom kecil di web untuk teks dan tombol hapus
                col_text, col_btn = st.columns([5, 1])
                with col_text:
                    st.write(f"📅 **{tgl_b}** | {nama_b} — **Rp {harga_b:,.0f}".replace(",", "."))
                with col_btn:
                    # Tombol hapus unik berdasarkan nomor baris asli di file Excel
                    if st.button("🗑️ Hapus", key=f"del_{row_id}"):
                        # Hapus baris di excel
                        wb.delete_rows(row_id)
                        wb.save(DB_FILE)
                        st.success("Data berhasil dihapus!")
                        st.rerun()
                        
            st.markdown("---")
            st.metric(label="Total Pengeluaran Anda Saat Ini", value=f"Rp {total_user:,.0f}".replace(",", "."))
            
            # Ekspor ulang hasil filter ke file untuk didownload jika diperlukan
            st.subheader("📂 Download Data Pribadi")
            df_download = pd.DataFrame([e[1:6] for e in user_entries], columns=["Hari", "Tanggal", "Bulan", "Nama Barang / Kebutuhan", "Harga (Rp)"])
            file_download_name = f"Pengeluaran_{st.session_state.username}.xlsx"
            df_download.to_excel(file_download_name, index=False)
            
            with open(file_download_name, "rb") as file:
                st.download_button(label="📥 Download Excel Pengeluaran Saya", data=file, file_name=file_download_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if os.path.exists(file_download_name):
                os.remove(file_download_name)
        else:
            st.info("Belum ada riwayat pengeluaran pada akun Anda.")
    except Exception as e:
        st.error("Gagal membaca atau memproses database.")
